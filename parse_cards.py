import openpyxl, re, json, sys

SRC = "/mnt/user-data/uploads/Крутагидон_база_карт.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

HEADERS = ["id","name","type","legend_subtype","cost","power","vp",
           "has_attack","attack_text","has_defense","defense_text",
           "postoyanka","activation","act_before","act_after",
           "full_text","chipsina_symbol","familiar_owner","notes","photo"]

def clean(v):
    return v.strip() if isinstance(v, str) else v

ws = wb["Карты"]
rows = []
for r in range(2, ws.max_row+1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, len(HEADERS)+1)]
    rid = clean(vals[0])
    if not rid:
        continue  # section divider / blank row
    if rid.endswith("_example"):
        continue  # template example row, not real data
    if not vals[2]:
        continue  # section-title row (no Type set), e.g. "Легенды и их виды"
    d = dict(zip(HEADERS, vals))
    d["row"] = r
    rows.append(d)

print(f"Всего реальных строк карт (без примеров/разделителей): {len(rows)}")

# ----- Известные тиражи из официального списка компонентов -----
KNOWN_QTY = {
    "start_znak": 30,
    "start_pshik": 15,
    "start_syrpal": 5,
    "start_hrenal": 1,
    "spec_vyal": 15,
    "spec_wild": 15,
}

QTY_PATTERNS = [
    (re.compile(r"в количестве двух|в количестве 2"), 2),
    (re.compile(r"единственная в колоде|в количестве одной"), 1),
    (re.compile(r"всего (\d+) шт"), None),  # captured group
]

def infer_qty(row):
    rid = row["id"]
    if rid in KNOWN_QTY:
        return KNOWN_QTY[rid], "known_component_count"
    notes = row.get("notes") or ""
    for pat, fixed in QTY_PATTERNS:
        m = pat.search(notes)
        if m:
            if fixed is not None:
                return fixed, "notes_pattern"
            else:
                return int(m.group(1)), "notes_explicit_number"
    # section-based defaults
    t = row["type"]
    if t in ("Легенда","Фамильяр","Беспредел","Мегабеспредел"):
        return 1, "default_unique_by_type"
    return None, "UNKNOWN"

unknown = []
for row in rows:
    qty, source = infer_qty(row)
    row["qty_in_deck"] = qty
    row["qty_source"] = source
    if qty is None:
        unknown.append(row)

print(f"Не удалось однозначно определить тираж: {len(unknown)} строк")
for row in unknown:
    t = row['type'] or '?'
    print(f"  строка {row['row']:>3}  id={row['id']:<20} тип={t:<12} заметки={ (row['notes'] or '')[:70] }")

with open("/home/claude/krutagidon/cards.json", "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

# ---------------------------------------------------------------------
# Жетоны дохлых колдунов (ЖДК)
# ---------------------------------------------------------------------
ws_zhdk = wb["Жетоны дохлых колдунов"]
ZHDK_HEADERS = ["id", "name", "effect_text", "postoyanka", "vp_penalty", "notes", "photo"]
zhdk_rows = []
for r in range(2, ws_zhdk.max_row + 1):
    vals = [ws_zhdk.cell(row=r, column=c).value for c in range(1, len(ZHDK_HEADERS) + 1)]
    rid = clean(vals[0])
    if not rid:
        continue
    d = dict(zip(ZHDK_HEADERS, vals))
    d["postoyanka"] = str(d.get("postoyanka") or "").strip().lower().startswith("да")
    try:
        d["vp_penalty"] = int(d.get("vp_penalty") or -3)
    except (TypeError, ValueError):
        d["vp_penalty"] = -3
    zhdk_rows.append(d)

with open("/home/claude/krutagidon/zhdk.json", "w", encoding="utf-8") as f:
    json.dump(zhdk_rows, f, ensure_ascii=False, indent=2)
print(f"ЖДК: {len(zhdk_rows)} жетонов сохранено в zhdk.json")

# ---------------------------------------------------------------------
# Жетоны колдунских свойств
# ---------------------------------------------------------------------
ws_svo = wb["Жетоны колдунских свойств"]
SVO_HEADERS = ["id", "name", "effect_text", "notes", "photo"]
svo_rows = []
for r in range(2, ws_svo.max_row + 1):
    vals = [ws_svo.cell(row=r, column=c).value for c in range(1, len(SVO_HEADERS) + 1)]
    rid = clean(vals[0])
    if not rid:
        continue
    svo_rows.append(dict(zip(SVO_HEADERS, vals)))

with open("/home/claude/krutagidon/svo.json", "w", encoding="utf-8") as f:
    json.dump(svo_rows, f, ensure_ascii=False, indent=2)
print(f"Колдунские свойства: {len(svo_rows)} шт сохранено в svo.json")

# ---------------------------------------------------------------------
# Планшеты колдунов (имя -> id фамильяра)
# ---------------------------------------------------------------------
ws_boards = wb["Планшеты колдунов"]
board_rows = []
for r in range(2, ws_boards.max_row + 1):
    vals = [ws_boards.cell(row=r, column=c).value for c in range(1, 4)]
    if not vals[0]:
        continue
    board_rows.append({"colduna_name": vals[0], "familiar_id": vals[1], "notes": vals[2]})

with open("/home/claude/krutagidon/boards.json", "w", encoding="utf-8") as f:
    json.dump(board_rows, f, ensure_ascii=False, indent=2)
print(f"Планшеты колдунов: {len(board_rows)} шт сохранено в boards.json")

# sanity totals by type
from collections import defaultdict
totals = defaultdict(int)
for row in rows:
    if row["qty_in_deck"]:
        totals[row["type"]] += row["qty_in_deck"]
print("\nСуммарный тираж по типам (где тираж известен):")
for t, n in sorted(totals.items()):
    print(f"  {t:<16} {n}")
